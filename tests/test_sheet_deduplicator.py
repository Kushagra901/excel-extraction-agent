import tempfile
from pathlib import Path

import openpyxl

from src.orchestrator import Supervisor
from src.utils.file_helpers import DEFAULT_CONFIG


def test_two_identical_sheets_deduplicated(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Full Name", "Email"])
    ws1.append(["Alice Smith", "alice@example.com"])
    ws1.append(["Bob Jones", "bob@example.com"])

    ws2 = wb.create_sheet(title="Copy of Sheet1")
    ws2.append(["Full Name", "Email"])
    ws2.append(["Alice Smith", "alice@example.com"])
    ws2.append(["Bob Jones", "bob@example.com"])

    excel_path = tmp_path / "identical_sheets.xlsx"
    wb.save(excel_path)

    out_dir = tmp_path / "output"
    supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
    ctx = supervisor.run(excel_path, out_dir)

    # Identical sheet records should be deduplicated (2 total records from Sheet1 only)
    assert len(ctx.records) == 2
    assert all(r.source_sheet == "Sheet1" for r in ctx.records)

    # Overlap report should log IDENTICAL action
    report = ctx.sheet_overlap_report.get("Sheet1 vs Copy of Sheet1")
    assert report is not None
    assert report["overlap_percentage"] == 100.0
    assert report["shared_row_count"] == 2
    assert "IDENTICAL" in report["action_taken"]


def test_two_sheets_partial_overlap_flagged(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "SheetA"
    ws1.append(["Full Name", "Email"])
    ws1.append(["Alice Smith", "alice@example.com"])
    ws1.append(["Bob Jones", "bob@example.com"])

    ws2 = wb.create_sheet(title="SheetB")
    ws2.append(["Full Name", "Email"])
    ws2.append(["Alice Smith", "alice@example.com"])
    ws2.append(["Charlie Brown", "charlie@example.com"])

    excel_path = tmp_path / "partial_overlap.xlsx"
    wb.save(excel_path)

    out_dir = tmp_path / "output"
    supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
    ctx = supervisor.run(excel_path, out_dir)

    # All records kept (2 from SheetA + 2 from SheetB = 4)
    assert len(ctx.records) == 4

    # Overlap report should log PARTIAL_OVERLAP
    report = ctx.sheet_overlap_report.get("SheetA vs SheetB")
    assert report is not None
    assert report["shared_row_count"] == 1
    assert "PARTIAL_OVERLAP" in report["action_taken"]
    assert any("Cross-sheet overlap detected" in i.message for i in ctx.issues)


def test_two_unique_sheets_no_changes(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.append(["Full Name", "Email"])
    ws1.append(["Alice Smith", "alice@example.com"])

    ws2 = wb.create_sheet(title="Products")
    ws2.append(["Product Name", "Price"])
    ws2.append(["Widget X", "$99.99"])

    excel_path = tmp_path / "unique_sheets.xlsx"
    wb.save(excel_path)

    out_dir = tmp_path / "output"
    supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
    ctx = supervisor.run(excel_path, out_dir)

    # Both records kept
    assert len(ctx.records) == 2
    report = ctx.sheet_overlap_report.get("Customers vs Products")
    assert report is not None
    assert report["shared_row_count"] == 0
    assert "UNIQUE" in report["action_taken"]
    assert len(ctx.sheet_diffs) == 1


def test_sheet_diff_same_key_modified_values(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Q1_Sales"
    ws1.append(["Email", "Full Name", "Amount"])
    ws1.append(["alice@example.com", "Alice Smith", "100.0"])

    ws2 = wb.create_sheet(title="Q2_Sales")
    ws2.append(["Email", "Full Name", "Amount"])
    ws2.append(["alice@example.com", "Alice Smith", "150.0"])

    excel_path = tmp_path / "modified_values.xlsx"
    wb.save(excel_path)

    out_dir = tmp_path / "output"
    supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
    ctx = supervisor.run(excel_path, out_dir)

    assert len(ctx.sheet_diffs) == 1
    diff = ctx.sheet_diffs[0]
    assert diff.sheet_a == "Q1_Sales"
    assert diff.sheet_b == "Q2_Sales"
    assert diff.common_record_count == 1
    assert len(diff.modified_records) == 1
    assert diff.modified_records[0]["field"] == "amount"
    assert diff.modified_records[0]["val_a"] == 100.0
    assert diff.modified_records[0]["val_b"] == 150.0

    report_text = (Path(ctx.output_dir) / "extraction_report.md").read_text(encoding="utf-8")
    assert "Sheet Comparison" in report_text
    assert "field 'amount' differs" in report_text

