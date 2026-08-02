import pytest
from pathlib import Path

from src.agents.excel_reader import ExcelReaderAgent

FIXTURE = Path(__file__).parent / "fixtures" / "messy_sample.xlsx"


def test_reads_all_sheets():
    reader = ExcelReaderAgent()
    profiles, grids = reader.read_workbook(FIXTURE)
    assert set(profiles.keys()) == {"Customers", "Orders", "Notes"}


def test_detects_header_row_past_title_and_blank_row():
    """Customers sheet has: merged title (row 1), blank row (row 2),
    real headers (row 3, 0-indexed 2). The detector must skip the first two."""
    reader = ExcelReaderAgent()
    profiles, grids = reader.read_workbook(FIXTURE)
    assert profiles["Customers"].header_row_index == 2


def test_merged_title_value_propagated_across_range():
    reader = ExcelReaderAgent()
    profiles, grids = reader.read_workbook(FIXTURE)
    grid = grids["Customers"]
    # A1:E1 merged -> every cell in row 0 should carry the title text.
    assert grid[0][0] == "Customer Export - Q2 2026"
    assert grid[0][4] == "Customer Export - Q2 2026"


def test_sparse_sheet_flagged():
    """`is_sparse` measures the fraction of empty cells across the sheet's
    *used range* -- it flags 'mostly blank cells inside a big range', not
    'few rows'. The Notes fixture has content in A1 and E20, so the used
    range is 20x5=100 cells with only 2 filled -> 98% empty -> sparse."""
    reader = ExcelReaderAgent(sparse_row_threshold=0.9)
    profiles, grids = reader.read_workbook(FIXTURE)
    assert profiles["Notes"].sparse_ratio >= 0.9
    assert profiles["Notes"].is_sparse


def test_oversized_file_rejected():
    reader = ExcelReaderAgent(max_file_size_mb=0.000001)
    with pytest.raises(ValueError, match="exceeds limit"):
        reader.read_workbook(FIXTURE)


def test_invalid_file_signature_rejected(tmp_path):
    fake_file = tmp_path / "fake.xlsx"
    fake_file.write_bytes(b"NOT A REAL XLSX ZIP FILE")
    reader = ExcelReaderAgent()
    with pytest.raises(ValueError, match="invalid file signature"):
        reader.read_workbook(fake_file)


def test_bad_zip_file_error_message():
    from unittest.mock import patch
    import zipfile
    reader = ExcelReaderAgent()
    with patch("openpyxl.load_workbook", side_effect=zipfile.BadZipFile("bad zip")):
        with pytest.raises(ValueError, match="corrupted or is not a valid .xlsx file"):
            reader.read_workbook(FIXTURE)


def test_password_protected_file_error_message():
    from unittest.mock import patch
    reader = ExcelReaderAgent()
    with patch("openpyxl.load_workbook", side_effect=Exception("File is encrypted and password protected")):
        with pytest.raises(ValueError, match="password-protected"):
            reader.read_workbook(FIXTURE)


def test_invalid_file_exception_error_message():
    from unittest.mock import patch
    import openpyxl.utils.exceptions
    reader = ExcelReaderAgent()
    with patch("openpyxl.load_workbook", side_effect=openpyxl.utils.exceptions.InvalidFileException("not valid")):
        with pytest.raises(ValueError, match="format is not supported"):
            reader.read_workbook(FIXTURE)


def test_detect_formulas_disabled():
    reader = ExcelReaderAgent(detect_formulas=False)
    profiles, grids = reader.read_workbook(FIXTURE)
    for profile in profiles.values():
        assert profile.has_formulas is False




