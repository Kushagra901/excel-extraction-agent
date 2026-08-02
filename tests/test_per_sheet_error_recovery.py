import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl

from src.agents.data_cleaner import DataCleaningAgent
from src.orchestrator import Supervisor
from src.utils.file_helpers import DEFAULT_CONFIG


def test_per_sheet_error_recovery_one_sheet_corrupted(tmp_path):
    # Create a workbook with 2 sheets: ValidSheet and CorruptedSheet
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "ValidSheet"
    ws1.append(["Full Name", "Email"])
    ws1.append(["Alice Smith", "alice@example.com"])

    ws2 = wb.create_sheet(title="CorruptedSheet")
    ws2.append(["Full Name", "Email"])
    ws2.append(["Bob Jones", "bob@example.com"])

    excel_path = tmp_path / "test_multi_sheet.xlsx"
    wb.save(excel_path)

    orig_clean_sheet = DataCleaningAgent.clean_sheet

    def mock_clean_sheet(self, grid, header_row_index, headers, schema_map, sheet_name, ctx):
        if sheet_name == "CorruptedSheet":
            raise RuntimeError("Corrupted sheet layout structure")
        return orig_clean_sheet(self, grid, header_row_index, headers, schema_map, sheet_name, ctx)

    with patch.object(DataCleaningAgent, "clean_sheet", side_effect=mock_clean_sheet, autospec=True):
        out_dir = tmp_path / "output"
        supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
        ctx = supervisor.run(excel_path, out_dir)

        # ValidSheet records should be successfully extracted
        assert len(ctx.records) == 1
        assert ctx.records[0].source_sheet == "ValidSheet"
        assert ctx.records[0].data.get("full_name") == "Alice Smith"

        # CorruptedSheet should have an ERROR issue recorded
        corrupted_issues = [
            i for i in ctx.issues
            if i.sheet == "CorruptedSheet" and i.severity.value == "error"
        ]
        assert len(corrupted_issues) == 1
        assert "failed during processing" in corrupted_issues[0].message
        assert "Corrupted sheet layout structure" in corrupted_issues[0].message

        run_dir = Path(ctx.output_dir)
        # Output formatter should have written all output deliverables
        assert (run_dir / "cleaned_data.csv").exists()
        assert (run_dir / "error_log.txt").exists()


def test_per_sheet_error_recovery_all_sheets_failed(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BadSheet"
    ws.append(["Header1", "Header2"])
    ws.append(["Val1", "Val2"])

    excel_path = tmp_path / "all_failed.xlsx"
    wb.save(excel_path)

    with patch.object(DataCleaningAgent, "clean_sheet", side_effect=ValueError("Global sheet failure")):
        out_dir = tmp_path / "output"
        supervisor = Supervisor(config=dict(DEFAULT_CONFIG))
        ctx = supervisor.run(excel_path, out_dir)

        # No records extracted, but pipeline completes and outputs are written
        assert len(ctx.records) == 0
        assert any("BadSheet" in i.sheet and i.severity.value == "error" for i in ctx.issues)
        run_dir = Path(ctx.output_dir)
        assert (run_dir / "cleaned_data.csv").exists()
        assert (run_dir / "extraction_report.md").exists()
