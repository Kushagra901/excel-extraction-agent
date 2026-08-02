"""
Excel Reader Agent.

Responsible for turning a raw .xlsx workbook into:
  1. A SheetProfile per sheet (structural facts: header row, merged cells,
     hidden rows/cols, formulas, sparsity).
  2. A raw grid (list of lists) per sheet, with merged-cell values
     propagated to every cell in the merged region so downstream agents
     never have to deal with None-filled merge artifacts.

This agent does NOT clean or interpret data -- it only describes structure
and hands over the raw grid. Interpretation is the Schema Mapping Agent's
and Data Cleaning Agent's job.
"""

from __future__ import annotations

import logging
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.utils.exceptions
from openpyxl.utils import range_boundaries

from src.core.models import IssueSeverity, MergedRegion, SheetProfile

logger = logging.getLogger("excel_agent.excel_reader")

MAX_HEADER_SEARCH_ROWS = 15


class ExcelReaderAgent:
    def __init__(
        self,
        sparse_row_threshold: float = 0.9,
        max_file_size_mb: float = 100,
        detect_formulas: bool = True,
        max_rows_per_sheet: int = 100000,
    ):
        self.sparse_row_threshold = sparse_row_threshold
        self.max_file_size_mb = max_file_size_mb
        self.detect_formulas = detect_formulas
        self.max_rows_per_sheet = max_rows_per_sheet

    @staticmethod
    def _validate_file_signature(path: Path) -> str:
        ext = path.suffix.lower()
        if ext in (".csv", ".tsv", ".txt"):
            return "csv"
        with open(path, "rb") as f:
            header = f.read(8)
        if header.startswith(b"PK\x03\x04"):
            return "xlsx"
        if header.startswith(b"\xd0\xcf\x11\xe0"):
            return "xls"
        if ext in (".xlsx", ".xls"):
            raise ValueError(
                "File does not appear to be a valid .xlsx or .xls file (invalid file signature). "
                "Ensure the file is not corrupted or renamed from another format."
            )
        raise ValueError("Unsupported file format. Only .xlsx, .xls, .csv, and .tsv files are supported.")

    def read_workbook(self, path: Path) -> tuple[dict[str, SheetProfile], dict[str, list[list[Any]]]]:
        """Returns (sheet_profiles, raw_grids) keyed by sheet name."""
        fmt = self._validate_file_signature(path)

        import os
        file_size_mb = os.path.getsize(path) / (1024 * 1024)
        if file_size_mb > self.max_file_size_mb:
            raise ValueError(
                f"File size {file_size_mb:.1f}MB exceeds limit of {self.max_file_size_mb}MB. "
                "Reduce the file size or increase max_file_size_mb in config."
            )

        if fmt == "csv":
            return self._read_csv_file(path)
        if fmt == "xls":
            return self._read_xls_workbook(path)

        # ------------------------------------------------------------------ #
        # Formula presence detection optimization:
        # Load formula workbook in read_only=True mode to prevent loading
        # full data structures twice into memory. Scan at most 100 rows per sheet,
        # and close the formula workbook immediately after inspection.
        # ------------------------------------------------------------------ #
        formula_presence: dict[str, bool] = {}
        if self.detect_formulas:
            try:
                wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
                for sheet_name in wb_formulas.sheetnames:
                    formula_presence[sheet_name] = self._detect_formulas(wb_formulas[sheet_name])
                wb_formulas.close()
            except Exception as exc:
                logger.warning("Failed to detect formulas: %s", exc)

        # ------------------------------------------------------------------ #
        # Note on data workbook loading:
        # Full mode (read_only=False) is required for wb_values because openpyxl's
        # ReadOnlyWorksheet does NOT parse or expose merged_cells.ranges or complete
        # row/column dimension structures. These structural details are required
        # downstream to un-merge values and detect hidden rows/cols.
        # ------------------------------------------------------------------ #
        try:
            wb_values = openpyxl.load_workbook(path, data_only=True)
        except zipfile.BadZipFile:
            raise ValueError("The file appears to be corrupted or is not a valid .xlsx file.")
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("The file format is not supported. Ensure it's a valid .xlsx file.")
        except Exception as exc:
            msg = str(exc).lower()
            if "encrypted" in msg or "password" in msg:
                raise ValueError(
                    "This file is password-protected. Please remove the password protection in Excel "
                    "(File > Info > Protect Workbook > Encrypt with Password > clear the password), "
                    "save it, and re-upload."
                ) from exc
            raise ValueError(f"Invalid or corrupted Excel file: {exc}") from exc

        profiles: dict[str, SheetProfile] = {}
        grids: dict[str, list[list[Any]]] = {}

        for sheet_name in wb_values.sheetnames:
            ws_values = wb_values[sheet_name]

            grid, trunc_note = self._build_grid(ws_values, sheet_name)
            merged_regions = self._extract_merged_regions(ws_values, grid, sheet_name)
            self._propagate_merged_values(grid, merged_regions, ws_values)

            hidden_rows = [
                r for r in range(1, ws_values.max_row + 1)
                if ws_values.row_dimensions[r].hidden
            ]
            hidden_cols = [
                letter for letter, dim in ws_values.column_dimensions.items()
                if dim.hidden
            ]
            has_formulas = formula_presence.get(sheet_name, False)

            header_row_idx, candidates = self._detect_header_row(grid)
            sparse_ratio = self._compute_sparse_ratio(grid)
            is_sparse = sparse_ratio >= self.sparse_row_threshold

            notes = []
            if trunc_note:
                notes.append(trunc_note)
            if header_row_idx is None:
                notes.append("Could not confidently detect a header row; "
                             "falling back to row 0 -- verify manually.")
            if is_sparse:
                notes.append("Sheet is {:.0f}% empty cells overall.".format(
                    sparse_ratio * 100))

            profiles[sheet_name] = SheetProfile(
                name=sheet_name,
                n_rows=len(grid),
                n_cols=max((len(r) for r in grid), default=0),
                header_row_index=header_row_idx if header_row_idx is not None else 0,
                candidate_header_rows=candidates,
                merged_regions=merged_regions,
                hidden_row_indices=hidden_rows,
                hidden_col_letters=hidden_cols,
                has_formulas=has_formulas,
                sparse_ratio=sparse_ratio,
                notes=notes,
            )
            grids[sheet_name] = grid

            logger.info(
                "Sheet '%s': %d rows x %d cols | header row (0-indexed): %s | "
                "merged regions: %d | hidden rows: %d | hidden cols: %d | formulas: %s",
                sheet_name, profiles[sheet_name].n_rows, profiles[sheet_name].n_cols,
                header_row_idx, len(merged_regions), len(hidden_rows), len(hidden_cols),
                has_formulas,
            )
            for note in notes:
                logger.warning("Sheet '%s': %s", sheet_name, note)

        wb_values.close()
        return profiles, grids

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #

    def _build_grid(self, ws, sheet_name: str = "") -> tuple[list[list[Any]], str | None]:
        grid: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            grid.append(list(row))
        return self._build_grid_from_rows(grid, sheet_name)

    @staticmethod
    def _extract_merged_regions(ws, grid: list[list[Any]], sheet_name: str) -> list[MergedRegion]:
        regions = []
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            anchor_value = None
            if 0 <= min_row - 1 < len(grid) and 0 <= min_col - 1 < len(grid[min_row - 1]):
                anchor_value = grid[min_row - 1][min_col - 1]
            regions.append(MergedRegion(sheet=sheet_name, range_str=str(merged_range),
                                         anchor_value=anchor_value))
        return regions

    @staticmethod
    def _propagate_merged_values(grid: list[list[Any]], regions: list[MergedRegion], ws) -> None:
        """Fill every cell in a merged region with the anchor's value so
        downstream code never has to special-case None-from-merge."""
        for region in regions:
            min_col, min_row, max_col, max_row = range_boundaries(region.range_str)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ri, ci = r - 1, c - 1
                    if 0 <= ri < len(grid) and 0 <= ci < len(grid[ri]):
                        grid[ri][ci] = region.anchor_value

    @staticmethod
    def _detect_formulas(ws_formulas, max_rows: int = 100) -> bool:
        for row in ws_formulas.iter_rows(max_row=max_rows):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
        return False

    def _detect_header_row(self, grid: list[list[Any]]) -> tuple[int | None, list[int]]:
        """
        Heuristic header-row scoring:
          - higher fraction of non-empty cells that are strings -> more header-like
          - higher fraction of unique values -> more header-like (headers repeat rarely)
          - fully-numeric rows score low (usually data, not headers)
        Only searches the first MAX_HEADER_SEARCH_ROWS rows -- headers deep
        below that are rare and usually indicate a report needing manual review.
        """
        best_idx, best_score = None, 0.0
        candidates: list[int] = []
        search_limit = min(MAX_HEADER_SEARCH_ROWS, len(grid))

        for idx in range(search_limit):
            row = grid[idx]
            non_empty = [c for c in row if c is not None and str(c).strip() != ""]
            if not non_empty:
                continue

            str_ratio = sum(1 for c in non_empty if isinstance(c, str)) / len(non_empty)
            unique_ratio = len(set(str(c).strip().lower() for c in non_empty)) / len(non_empty)
            fill_ratio = len(non_empty) / max(len(row), 1)

            score = (str_ratio * 0.5) + (unique_ratio * 0.3) + (fill_ratio * 0.2)

            if score >= 0.55:
                candidates.append(idx)
            if score > best_score:
                best_score, best_idx = score, idx

        if best_score < 0.4:
            return None, candidates
        return best_idx, candidates

    @staticmethod
    def _compute_sparse_ratio(grid: list[list[Any]]) -> float:
        total = sum(len(row) for row in grid)
        if total == 0:
            return 1.0
        empty = sum(
            1 for row in grid for c in row if c is None or str(c).strip() == ""
        )
        return empty / total

    def _read_xls_workbook(self, path: Path) -> tuple[dict[str, SheetProfile], dict[str, list[list[Any]]]]:
        import pandas as pd

        try:
            dict_of_dfs = pd.read_excel(path, sheet_name=None, engine="xlrd", header=None)
        except Exception as exc:
            raise ValueError(f"Invalid or corrupted legacy .xls Excel file: {exc}") from exc

        profiles: dict[str, SheetProfile] = {}
        grids: dict[str, list[list[Any]]] = {}

        for sheet_name, df in dict_of_dfs.items():
            raw_grid: list[list[Any]] = []
            for row in df.itertuples(index=False):
                row_vals = [
                    (None if pd.isna(v) or str(v).strip() in ("", "nan", "NaN", "None") else v)
                    for v in row
                ]
                raw_grid.append(row_vals)

            grid, trunc_note = self._build_grid_from_rows(raw_grid, sheet_name)

            header_row_idx, candidates = self._detect_header_row(grid)
            sparse_ratio = self._compute_sparse_ratio(grid)
            is_sparse = sparse_ratio >= self.sparse_row_threshold

            notes = [
                "Legacy .xls format — merged cell detection not available",
                "Formula detection not available for legacy .xls format",
            ]
            if trunc_note:
                notes.insert(0, trunc_note)
            if header_row_idx is None:
                notes.append("Could not confidently detect a header row; falling back to row 0 -- verify manually.")
            if is_sparse:
                notes.append("Sheet is {:.0f}% empty cells overall.".format(sparse_ratio * 100))

            profiles[sheet_name] = SheetProfile(
                name=sheet_name,
                n_rows=len(grid),
                n_cols=max((len(r) for r in grid), default=0),
                header_row_index=header_row_idx if header_row_idx is not None else 0,
                candidate_header_rows=candidates,
                merged_regions=[],
                hidden_row_indices=[],
                hidden_col_letters=[],
                has_formulas=False,
                sparse_ratio=sparse_ratio,
                notes=notes,
            )
            grids[sheet_name] = grid

            logger.info(
                "Sheet '%s' (.xls): %d rows x %d cols | header row (0-indexed): %s",
                sheet_name, profiles[sheet_name].n_rows, profiles[sheet_name].n_cols, header_row_idx,
            )

        return profiles, grids

    def _read_csv_file(self, path: Path) -> tuple[dict[str, SheetProfile], dict[str, list[list[Any]]]]:
        import csv
        import chardet
        import pandas as pd

        with open(path, "rb") as f:
            raw_bytes = f.read(10240)

        detected_enc = chardet.detect(raw_bytes).get("encoding")
        encoding = detected_enc if detected_enc else "utf-8"

        try:
            sample_text = raw_bytes.decode(encoding, errors="replace")
        except Exception:
            sample_text = raw_bytes.decode("utf-8", errors="replace")
            encoding = "utf-8"

        sample_lines = "\n".join(sample_text.splitlines()[:5])
        delimiter = ","
        if sample_lines.strip():
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample_lines, delimiters=[",", ";", "\t", "|"])
                delimiter = dialect.delimiter
            except Exception:
                if "\t" in sample_lines:
                    delimiter = "\t"
                elif ";" in sample_lines:
                    delimiter = ";"
                elif "|" in sample_lines:
                    delimiter = "|"
                else:
                    delimiter = ","

        try:
            df = pd.read_csv(path, encoding=encoding, sep=delimiter, header=None, dtype=str)
        except Exception as exc:
            raise ValueError(f"Failed to parse CSV file: {exc}") from exc

        raw_grid: list[list[Any]] = []
        for row in df.itertuples(index=False):
            row_vals = [
                (None if pd.isna(v) or str(v).strip() in ("", "nan", "NaN", "None") else v)
                for v in row
            ]
            raw_grid.append(row_vals)

        sheet_name = "csv_data"
        grid, trunc_note = self._build_grid_from_rows(raw_grid, sheet_name)

        header_row_idx, candidates = self._detect_header_row(grid)
        sparse_ratio = self._compute_sparse_ratio(grid)
        is_sparse = sparse_ratio >= self.sparse_row_threshold

        delim_display = {",": ",", ";": ";", "\t": "TAB", "|": "|"}.get(delimiter, delimiter)
        notes = [
            f"CSV import — delimiter: '{delim_display}', encoding: '{encoding}'"
        ]
        if trunc_note:
            notes.insert(0, trunc_note)
        if header_row_idx is None:
            notes.append("Could not confidently detect a header row; falling back to row 0 -- verify manually.")
        if is_sparse:
            notes.append("Sheet is {:.0f}% empty cells overall.".format(sparse_ratio * 100))

        profiles = {
            sheet_name: SheetProfile(
                name=sheet_name,
                n_rows=len(grid),
                n_cols=max((len(r) for r in grid), default=0),
                header_row_index=header_row_idx if header_row_idx is not None else 0,
                candidate_header_rows=candidates,
                merged_regions=[],
                hidden_row_indices=[],
                hidden_col_letters=[],
                has_formulas=False,
                sparse_ratio=sparse_ratio,
                notes=notes,
            )
        }
        grids = {sheet_name: grid}

        logger.info(
            "CSV Sheet '%s': %d rows x %d cols | header row (0-indexed): %s | delimiter: '%s' | encoding: '%s'",
            sheet_name, profiles[sheet_name].n_rows, profiles[sheet_name].n_cols, header_row_idx,
            delim_display, encoding,
        )

        return profiles, grids

    def _build_grid_from_rows(self, grid: list[list[Any]], sheet_name: str = "") -> tuple[list[list[Any]], str | None]:
        trunc_note = None
        if self.max_rows_per_sheet > 0 and len(grid) > self.max_rows_per_sheet:
            actual = len(grid)
            limit = self.max_rows_per_sheet
            grid = grid[:limit]
            logger.warning(
                "Sheet '%s' has %d rows, truncated to %d. Set max_rows_per_sheet in config to increase.",
                sheet_name, actual, limit,
            )
            trunc_note = (
                f"Sheet '{sheet_name}' has {actual} rows, truncated to {limit}. "
                f"Set max_rows_per_sheet in config to increase."
            )
        return grid, trunc_note
